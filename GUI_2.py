# -*- coding: utf-8 -*-
"""
Created on Wed Apr 17 14:14:28 2024

@author: KS43772
"""

import tkinter as tk
from tkinter import filedialog,messagebox  
from ttkthemes import ThemedTk
import geopandas as gpd
from openpyxl import load_workbook
from datetime import datetime

# Global variables to store paths
geopackage_path = None
excel_path = None


# Function to process GeoPackage and Excel files
def process_files(geopackage_path, excel_path):
    # Ensure all paths are valid
    if not geopackage_path or not excel_path:
        print("GeoPackage and Excel file paths must be provided.")
        return
    
    # Load the GeoPackage file
    desired_layer = 'Tree_planting_monitoring_26c65948_4e4f_46a9_8e97_c6096663dd6e'
    gdf = gpd.read_file(geopackage_path, layer=desired_layer)
    print("GeoPackage file loaded successfully.")
    
    # Load the Excel file
    existing_wb = load_workbook(excel_path)
    print("Excel file loaded successfully.")

    # Example of updating the existing Excel file
    plot1_sheet = existing_wb['Plot 1']
    dbf_table = gdf.drop(columns=['geometry'])

    # Assuming the survey date is in the 'Date' column in the GeoPackage data
    survey_date = dbf_table['Date'].unique()
    date_obj = datetime.strptime(survey_date[0], "%Y-%m-%d")
    formatted_survey_date = date_obj.strftime("%d %B %Y")

    # Insert the survey date in the specified cell
    plot1_sheet.cell(row=3, column=21).value = formatted_survey_date

    # Define a dictionary mapping species names to column indices in Excel
    species_columns_map = {
        'Alder': 2,
        'Ash': 3,
        'Aspen': 4,
        'Birch': 5,
        'Blackthorn': 6,
        'Cherry': 7,
        'Elder': 8,
        'Hawthorn': 9,
        'Hazel': 10,
        'Holly': 11,
        'Juniper': 12,
        'Oak': 13,
        'Rowan': 14,
        'Scots Pine': 15,
        'Willow': 16
    }

    # Write the species names to row 9
    for species, column_index in species_columns_map.items():
        plot1_sheet.cell(row=9, column=column_index).value = species
    
    # Populate data based on the GeoPackage information
    next_row_index = 10
    for species, column_index in species_columns_map.items():
        species_entries = dbf_table.loc[dbf_table['Species'] == species]

        for _, entry in species_entries.iterrows():
            # Get height, damage, dead, health, etc.
            height = entry.get('Height', None)
            damage = entry.get('Damaged', None)  # Assuming it's a valid column
            dead = entry.get('Dead', None)  # Assuming it's a valid column
            health = entry.get('Growth_Hea', None)  # Assuming it's a valid column

            # Update the cells in the corresponding row and column
            plot1_sheet.cell(row=next_row_index, column=column_index).value = height
            plot1_sheet.cell(row=next_row_index, column=17).value = damage
            plot1_sheet.cell(row=next_row_index, column=18).value = dead
            plot1_sheet.cell(row=next_row_index, column=21).value = health

            # Move to the next row
            next_row_index += 1

    # Save the changes to the existing Excel file
    try:
        existing_wb.save(excel_path)
        print("Updated Excel file saved.")
        # Show success message
        messagebox.showinfo("Processing Complete", "The processing has been completed successfully.")
    except Exception as e:
        print(f"Error saving Excel: {e}")
        messagebox.showerror("Error", f"Failed to save Excel: {e}")
    


# Function to select GeoPackage file
def set_geopackage_path():
    global geopackage_path
    filepath = filedialog.askopenfilename(filetypes=[("GeoPackage files", "*.gpkg")])
    if filepath:
        geopackage_path = filepath


# Function to select Excel file
def set_excel_path():
    global excel_path
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filepath:
        excel_path = filepath


# Function to run processing code when OK button is clicked
def run_process():
    if geopackage_path and excel_path:
        process_files(geopackage_path, excel_path)
    else:
        print("Please select a GeoPackage and Excel file before running.")


# Create the main GUI window with ThemedTk
root = ThemedTk(theme="radiance")
root.title("Load GeoPackage and Excel")
root.geometry("400x200")

# Create buttons with larger size and padding for spacing
select_geopackage_button = tk.Button(
    root, 
    text="Select GeoPackage", 
    command=set_geopackage_path,
    width=20,  # Explicitly set the width
    height=2,  # Explicitly set the height
)
select_geopackage_button.pack(pady=10)  # 10-pixel padding between widgets

select_excel_button = tk.Button(
    root, 
    text="Select Excel", 
    command=set_excel_path,
    width=20,  # Set the same size for consistency
    height=2,  # Make buttons taller
)
select_excel_button.pack(pady=10)

# Create an OK button to run the process
ok_button = tk.Button(
    root, 
    text="OK", 
    command=run_process,
    width=20,  # Ensure uniformity in button size
    height=2,
)
ok_button.pack(pady=10)  # Add space between buttons

# Start the Tkinter main loop
root.mainloop()